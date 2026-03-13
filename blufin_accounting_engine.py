#!/usr/bin/env python3

import argparse
import csv
import json
import re
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


VALID_TYPES = {"income", "expense", "asset", "liability", "equity"}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")


def parse_amount(x):
    s = str(x).replace(",", "").replace('"', "").strip()
    if s == "":
        return 0.0
    return float(s)


def get_last_nonempty_cell(row):
    for cell in reversed(row):
        val = str(cell).strip()
        if val != "":
            return val
    return ""


def parse_statement_header(stmt_path):
    with open(stmt_path, "r", encoding="utf-8-sig") as f:
        lines = f.read().splitlines()

    tx_start_idx = None
    for i, line in enumerate(lines):
        if line.startswith("Date,Description,Amount,Running Bal."):
            tx_start_idx = i
            break

    if tx_start_idx is None:
        raise ValueError("Transaction section not found in statement CSV.")

    header = {
        "beginning_balance_label": "Beginning balance",
        "beginning_balance": 0.0,
        "beginning_balance_date": None,
        "total_credits": 0.0,
        "total_debits": 0.0,
        "ending_balance_label": "Ending balance",
        "ending_balance": 0.0,
        "ending_balance_date": None,
    }

    for line in lines[:tx_start_idx]:
        if not line.strip():
            continue

        row = next(csv.reader([line]))
        if not row:
            continue

        label = str(row[0]).strip() if len(row) > 0 else ""
        if not label or label.upper() == "DESCRIPTION":
            continue

        value = get_last_nonempty_cell(row)
        label_u = label.upper()
        date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", label)

        if "BEGINNING BALANCE AS OF" in label_u:
            header["beginning_balance_label"] = label
            header["beginning_balance"] = parse_amount(value)
            if date_match:
                header["beginning_balance_date"] = pd.to_datetime(date_match.group(1), format="%m/%d/%Y")

        elif "TOTAL CREDITS" in label_u:
            header["total_credits"] = parse_amount(value)

        elif "TOTAL DEBITS" in label_u:
            header["total_debits"] = parse_amount(value)

        elif "ENDING BALANCE AS OF" in label_u:
            header["ending_balance_label"] = label
            header["ending_balance"] = parse_amount(value)
            if date_match:
                header["ending_balance_date"] = pd.to_datetime(date_match.group(1), format="%m/%d/%Y")

    return header


def get_statement_period(stmt_header):
    begin_date = stmt_header.get("beginning_balance_date")
    end_date = stmt_header.get("ending_balance_date")

    if begin_date is None or end_date is None:
        raise ValueError("Could not determine statement period from statement header.")

    period_start = begin_date + pd.Timedelta(days=1)
    period_end = end_date
    return period_start, period_end


def load_transactions(xlsx_path):
    df = pd.read_excel(xlsx_path, sheet_name="Raw_GL")

    required = ["date", "description", "amount", "category"]
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Missing column in Raw_GL: {col}")

    for col in ["counterparty", "account", "memo"]:
        if col not in df.columns:
            df[col] = ""

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    if df["date"].isna().any():
        bad = df[df["date"].isna()]
        raise ValueError(
            "Invalid dates in Raw_GL:\n"
            + bad[["date", "description", "amount"]].to_string(index=False)
        )

    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)
    df["description"] = df["description"].fillna("").astype(str)
    df["category"] = df["category"].fillna("").astype(str)
    df["counterparty"] = df["counterparty"].fillna("").astype(str)
    df["account"] = df["account"].fillna("").astype(str)
    df["memo"] = df["memo"].fillna("").astype(str)

    return df.sort_values("date").reset_index(drop=True)


def load_account_map(path):
    with open(path, "r", encoding="utf-8") as f:
        mapping = json.load(f)

    if "categories" not in mapping:
        raise ValueError("account_map.json must contain a 'categories' object.")

    for category, cfg in mapping["categories"].items():
        stype = cfg.get("statement_type")
        if stype not in VALID_TYPES:
            raise ValueError(
                f"Category '{category}' has invalid statement_type '{stype}'. "
                f"Expected one of {sorted(VALID_TYPES)}."
            )

    return mapping


def apply_types(df, mapping):
    df = df.copy()

    unknown = sorted(set(df["category"]) - set(mapping["categories"]))
    if unknown:
        raise ValueError(
            "These categories are missing from account_map.json: "
            + ", ".join(unknown)
        )

    df["type"] = df["category"].map(lambda c: mapping["categories"][c]["statement_type"])
    df["report_name"] = df["category"].map(
        lambda c: mapping["categories"][c].get("report_name", c)
    )
    return df


def filter_asof(df, asof=None):
    if not asof:
        return df.copy()
    cutoff = pd.Timestamp(asof)
    return df[df["date"] <= cutoff].copy()


def quarter_bounds(dt):
    quarter = ((dt.month - 1) // 3) + 1
    start_month = 3 * (quarter - 1) + 1
    start = pd.Timestamp(year=dt.year, month=start_month, day=1)
    if quarter == 4:
        end = pd.Timestamp(year=dt.year + 1, month=1, day=1) - pd.Timedelta(days=1)
    else:
        end = pd.Timestamp(year=dt.year, month=start_month + 3, day=1) - pd.Timedelta(days=1)
    return quarter, start, end


def previous_quarter_bounds(dt):
    q, start, _ = quarter_bounds(dt)
    if q == 1:
        prev_end = pd.Timestamp(year=dt.year, month=1, day=1) - pd.Timedelta(days=1)
    else:
        prev_end = start - pd.Timedelta(days=1)
    _, prev_start, prev_q_end = quarter_bounds(prev_end)
    return prev_start, prev_q_end


def compute_period_pnl(df, start_date=None, end_date=None):
    work = df.copy()

    if start_date is not None:
        work = work[work["date"] >= pd.Timestamp(start_date)]
    if end_date is not None:
        work = work[work["date"] <= pd.Timestamp(end_date)]

    income_df = work[work["type"] == "income"].copy()
    expense_df = work[work["type"] == "expense"].copy()

    income_by_line = income_df.groupby("report_name", dropna=False)["amount"].sum().sort_index()
    expense_by_line = (-expense_df.groupby("report_name", dropna=False)["amount"].sum()).sort_index()

    income = float(income_df["amount"].sum())
    expenses = float(-expense_df["amount"].sum())
    net_income = income - expenses

    return {
        "income": income,
        "expenses": expenses,
        "net_income": net_income,
        "income_by_line": income_by_line,
        "expense_by_line": expense_by_line,
    }


def compute_cashflow(df, stmt_header, asof=None):
    work = filter_asof(df, asof)

    period_start, period_end = get_statement_period(stmt_header)

    current_period = work[
        (work["date"] >= period_start) &
        (work["date"] <= period_end)
    ].copy()

    income_cash = float(current_period.loc[current_period["type"] == "income", "amount"].sum())
    expense_cash = float(current_period.loc[current_period["type"] == "expense", "amount"].sum())  # negative

    operating = income_cash + expense_cash
    investing = 0.0
    financing = 0.0

    return {
        "operating": operating,
        "investing": investing,
        "financing": financing,
        "net_cash_change": operating + investing + financing,
        "cash_in": float(stmt_header.get("total_credits", 0.0)),
        "cash_out": float(stmt_header.get("total_debits", 0.0)),
    }


def compute_balance(df, stmt_header, asof=None):
    work = filter_asof(df, asof)

    period_start, period_end = get_statement_period(stmt_header)
    pnl = compute_period_pnl(work, start_date=period_start, end_date=period_end)

    current_period_earnings = float(pnl["net_income"])
    retained_earnings = float(stmt_header.get("beginning_balance", 0.0))
    checking_balance = float(stmt_header.get("ending_balance", 0.0))

    # Pull actual balance-sheet activity from Raw_GL
    asset_df = work[work["type"] == "asset"].copy()
    liability_df = work[work["type"] == "liability"].copy()
    equity_df = work[work["type"] == "equity"].copy()

    # IMPORTANT:
    # Opening retained earnings is already taken from the statement header,
    # so exclude Retained Earnings rows from ledger equity activity to avoid double count.
    equity_df = equity_df[equity_df["report_name"] != "Retained Earnings"].copy()

    # Checking comes from the bank statement, not from ledger rows
    assets_by_line = pd.Series({"Checking": checking_balance}, dtype=float)

    liabilities_by_line = (
        liability_df.groupby("report_name", dropna=False)["amount"]
        .sum()
        .sort_index()
        if not liability_df.empty
        else pd.Series(dtype=float)
    )

    equity_activity_by_line = (
        equity_df.groupby("report_name", dropna=False)["amount"]
        .sum()
        .sort_index()
        if not equity_df.empty
        else pd.Series(dtype=float)
    )

    liabilities = float(liabilities_by_line.sum()) if not liabilities_by_line.empty else 0.0
    equity_activity = float(equity_activity_by_line.sum()) if not equity_activity_by_line.empty else 0.0

    assets = checking_balance
    total_equity = retained_earnings + current_period_earnings + equity_activity

    return {
        "assets": float(assets),
        "liabilities": float(liabilities),
        "equity_activity": float(equity_activity),
        "retained_earnings": float(retained_earnings),
        "current_year_earnings": float(current_period_earnings),
        "total_equity": float(total_equity),
        "balance_check": float(assets - liabilities - total_equity),
        "assets_by_line": assets_by_line,
        "liabilities_by_line": liabilities_by_line,
        "equity_activity_by_line": equity_activity_by_line,
    }

    liabilities = float(liabilities_by_line.sum()) if not liabilities_by_line.empty else 0.0
    equity_activity = float(equity_activity_by_line.sum()) if not equity_activity_by_line.empty else 0.0

    assets = checking_balance
    total_equity = retained_earnings + current_period_earnings + equity_activity

    return {
        "assets": float(assets),
        "liabilities": float(liabilities),
        "equity_activity": float(equity_activity),
        "retained_earnings": float(retained_earnings),
        "current_year_earnings": float(current_period_earnings),
        "total_equity": float(total_equity),
        "balance_check": float(assets - liabilities - total_equity),
        "assets_by_line": assets_by_line,
        "liabilities_by_line": liabilities_by_line,
        "equity_activity_by_line": equity_activity_by_line,
    }


def build_pnl_dataframe(pnl, period_label):
    rows = []
    for line, amt in pnl["income_by_line"].items():
        rows.append({"Section": "Income", "Line Item": line, period_label: float(amt)})
    rows.append({"Section": "Income", "Line Item": "Total Income", period_label: pnl["income"]})

    for line, amt in pnl["expense_by_line"].items():
        rows.append({"Section": "Expenses", "Line Item": line, period_label: float(amt)})
    rows.append({"Section": "Expenses", "Line Item": "Total Expenses", period_label: pnl["expenses"]})
    rows.append({"Section": "Summary", "Line Item": "Net Income", period_label: pnl["net_income"]})
    return pd.DataFrame(rows)


def build_balance_dataframe(balance, asof_label):
    rows = []

    for line, amt in balance["assets_by_line"].items():
        rows.append({"Section": "Assets", "Line Item": line, asof_label: float(amt)})
    rows.append({"Section": "Assets", "Line Item": "Total Assets", asof_label: balance["assets"]})

    for line, amt in balance["liabilities_by_line"].items():
        rows.append({"Section": "Liabilities", "Line Item": line, asof_label: float(amt)})
    rows.append({"Section": "Liabilities", "Line Item": "Total Liabilities", asof_label: balance["liabilities"]})

    for line, amt in balance["equity_activity_by_line"].items():
        rows.append({"Section": "Equity", "Line Item": line, asof_label: float(amt)})

    rows.append({"Section": "Equity", "Line Item": "Retained Earnings", asof_label: balance["retained_earnings"]})
    rows.append({"Section": "Equity", "Line Item": "Current Period Earnings", asof_label: balance["current_year_earnings"]})
    rows.append({"Section": "Equity", "Line Item": "Total Equity", asof_label: balance["total_equity"]})
    rows.append({"Section": "Check", "Line Item": "Balance Check", asof_label: balance["balance_check"]})

    return pd.DataFrame(rows)


def build_cashflow_dataframe(cf, period_label):
    return pd.DataFrame(
        [
            {"Section": "Operating", "Line Item": "Net Cash from Operating", period_label: cf["operating"]},
            {"Section": "Investing", "Line Item": "Net Cash from Investing", period_label: cf["investing"]},
            {"Section": "Financing", "Line Item": "Net Cash from Financing", period_label: cf["financing"]},
            {"Section": "Summary", "Line Item": "Cash In", period_label: cf["cash_in"]},
            {"Section": "Summary", "Line Item": "Cash Out", period_label: cf["cash_out"]},
            {"Section": "Summary", "Line Item": "Net Cash Change", period_label: cf["net_cash_change"]},
        ]
    )


def income_by_customer(df, start_date=None, end_date=None):
    work = df.copy()
    if start_date is not None:
        work = work[work["date"] >= pd.Timestamp(start_date)]
    if end_date is not None:
        work = work[work["date"] <= pd.Timestamp(end_date)]

    inc = work[work["type"] == "income"].copy()
    if inc.empty:
        return pd.DataFrame({"Customer": ["Unspecified"], "Income": [0.0]})

    inc["Customer"] = inc["counterparty"].fillna("").replace("", "Unspecified")
    out = inc.groupby("Customer", dropna=False)["amount"].sum().reset_index()
    out.columns = ["Customer", "Income"]
    return out.sort_values("Income", ascending=False).reset_index(drop=True)


def rolling_monthly_income_expense(df):
    work = df.copy()
    if work.empty:
        return pd.DataFrame(columns=["Month", "Income", "Expense", "Rolling Income", "Rolling Expense"])

    work["Month"] = work["date"].dt.to_period("M").dt.to_timestamp()
    income = work[work["type"] == "income"].groupby("Month")["amount"].sum()
    expense = -work[work["type"] == "expense"].groupby("Month")["amount"].sum()

    all_months = pd.date_range(work["Month"].min(), work["Month"].max(), freq="MS")

    monthly = pd.DataFrame(index=all_months)
    monthly.index.name = "Month"
    monthly["Income"] = income.reindex(all_months, fill_value=0.0)
    monthly["Expense"] = expense.reindex(all_months, fill_value=0.0)
    monthly["Rolling Income"] = monthly["Income"].cumsum()
    monthly["Rolling Expense"] = monthly["Expense"].cumsum()

    return monthly.reset_index()


def pnl_quarter_comparison(df, asof_date):
    current_q, current_start, current_end = quarter_bounds(asof_date)
    prev_start, prev_end = previous_quarter_bounds(asof_date)

    curr = compute_period_pnl(df, current_start, min(current_end, asof_date))
    prev = compute_period_pnl(df, prev_start, prev_end)

    lines = sorted(
        set(curr["income_by_line"].index)
        | set(curr["expense_by_line"].index)
        | set(prev["income_by_line"].index)
        | set(prev["expense_by_line"].index)
    )

    rows = []
    for line in lines:
        if line in curr["income_by_line"].index or line in prev["income_by_line"].index:
            section = "Income"
            curr_amt = float(curr["income_by_line"].get(line, 0.0))
            prev_amt = float(prev["income_by_line"].get(line, 0.0))
            pct_change = None if prev_amt == 0 else (curr_amt - prev_amt) / prev_amt
        else:
            section = "Expense"
            curr_amt = float(curr["expense_by_line"].get(line, 0.0))
            prev_amt = float(prev["expense_by_line"].get(line, 0.0))
            pct_change = None if prev_amt == 0 else (curr_amt - prev_amt) / prev_amt

        rows.append(
            {
                "Section": section,
                "Line Item": line,
                "Current Quarter": curr_amt,
                "Previous Quarter": prev_amt,
                "Pct Change": pct_change,
            }
        )

    rows.append(
        {
            "Section": "Income",
            "Line Item": "Total Income",
            "Current Quarter": curr["income"],
            "Previous Quarter": prev["income"],
            "Pct Change": None if prev["income"] == 0 else (curr["income"] - prev["income"]) / prev["income"],
        }
    )
    rows.append(
        {
            "Section": "Expenses",
            "Line Item": "Total Expenses",
            "Current Quarter": curr["expenses"],
            "Previous Quarter": prev["expenses"],
            "Pct Change": None if prev["expenses"] == 0 else (curr["expenses"] - prev["expenses"]) / prev["expenses"],
        }
    )
    rows.append(
        {
            "Section": "Summary",
            "Line Item": "Net Income",
            "Current Quarter": curr["net_income"],
            "Previous Quarter": prev["net_income"],
            "Pct Change": None if prev["net_income"] == 0 else (curr["net_income"] - prev["net_income"]) / prev["net_income"],
        }
    )

    return pd.DataFrame(rows), current_start, min(current_end, asof_date), prev_start, prev_end


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 32)


def write_dataframe(ws, df, start_row=1, start_col=1, currency_cols=None, pct_cols=None):
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []

    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=col)

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            ws.cell(row=i, column=j, value=row[col])

    style_header(ws, start_row)

    for j, col in enumerate(df.columns, start=start_col):
        if col in currency_cols:
            for i in range(start_row + 1, start_row + 1 + len(df)):
                ws.cell(row=i, column=j).number_format = '#,##0.00'
        if col in pct_cols:
            for i in range(start_row + 1, start_row + 1 + len(df)):
                ws.cell(row=i, column=j).number_format = '0.0%'


def preserve_raw_gl_sheet(target_wb, source_wb):
    if "Raw_GL" in target_wb.sheetnames:
        target_wb.remove(target_wb["Raw_GL"])

    source_ws = source_wb["Raw_GL"]
    new_ws = target_wb.create_sheet("Raw_GL", 0)

    for row in source_ws.iter_rows(values_only=True):
        new_ws.append(list(row))


def build_workbook(df, stmt_header, ledger_path, output_path, asof=None):
    work = filter_asof(df, asof)
    if work.empty:
        raise ValueError("No transactions found for selected reporting range.")

    asof_date = work["date"].max()
    period_start, period_end = get_statement_period(stmt_header)

    pnl = compute_period_pnl(work, period_start, period_end)
    balance = compute_balance(work, stmt_header, asof_date)
    cash = compute_cashflow(work, stmt_header, asof_date)

    pnl_df = build_pnl_dataframe(pnl, f"{period_start.date()} to {period_end.date()}")
    bs_df = build_balance_dataframe(balance, f"As of {asof_date.date()}")
    cf_df = build_cashflow_dataframe(cash, f"Through {period_end.date()}")

    customer_df = income_by_customer(work, period_start, period_end)
    rolling_df = rolling_monthly_income_expense(work)
    qcmp_df, cur_q_start, cur_q_end, prev_q_start, prev_q_end = pnl_quarter_comparison(work, asof_date)

    source_wb = openpyxl.load_workbook(ledger_path)
    if "Raw_GL" not in source_wb.sheetnames:
        raise ValueError("Workbook is missing Raw_GL sheet.")

    wb = openpyxl.Workbook()
    preserve_raw_gl_sheet(wb, source_wb)

    # remove default extra sheets except Raw_GL
    for ws_name in list(wb.sheetnames):
        if ws_name != "Raw_GL":
            wb.remove(wb[ws_name])

    ws_summary = wb.create_sheet("Summary")
    ws_pnl = wb.create_sheet("P&L")
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_cf = wb.create_sheet("Cash Flow")

    ws_summary["A1"] = "Blufin Labs Financial Summary"
    ws_summary["A2"] = f"As of {asof_date.date()}"
    ws_summary["A1"].font = Font(size=14, bold=True)

    summary_header_df = pd.DataFrame(
        [
            {"Item": stmt_header.get("beginning_balance_label", "Beginning balance"), "Amount": stmt_header.get("beginning_balance", 0.0)},
            {"Item": "Total credits", "Amount": stmt_header.get("total_credits", 0.0)},
            {"Item": "Total debits", "Amount": stmt_header.get("total_debits", 0.0)},
            {"Item": stmt_header.get("ending_balance_label", "Ending balance"), "Amount": stmt_header.get("ending_balance", 0.0)},
        ]
    )
    write_dataframe(ws_summary, summary_header_df, start_row=4, start_col=1, currency_cols=["Amount"])

    write_dataframe(ws_summary, customer_df, start_row=10, start_col=1, currency_cols=["Income"])
    pie = PieChart()
    pie.title = "Income vs Customer"
    labels = Reference(ws_summary, min_col=1, min_row=11, max_row=10 + len(customer_df))
    data = Reference(ws_summary, min_col=2, min_row=10, max_row=10 + len(customer_df))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 10
    ws_summary.add_chart(pie, "E10")

    write_dataframe(
        ws_summary,
        rolling_df,
        start_row=10,
        start_col=12,
        currency_cols=["Income", "Expense", "Rolling Income", "Rolling Expense"],
    )
    for r in range(11, 11 + len(rolling_df)):
        ws_summary.cell(row=r, column=12).number_format = "yyyy-mm"

    line = LineChart()
    line.title = "Rolling Income & Expense Totals"
    line.y_axis.title = "Amount"
    line.x_axis.title = "Month"
    data = Reference(ws_summary, min_col=15, max_col=16, min_row=10, max_row=10 + len(rolling_df))
    cats = Reference(ws_summary, min_col=12, min_row=11, max_row=10 + len(rolling_df))
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.height = 8
    line.width = 14
    ws_summary.add_chart(line, "Q10")

    ws_summary["A28"] = (
        f"Quarter Comparison: Current ({cur_q_start.date()} to {cur_q_end.date()}) "
        f"vs Previous ({prev_q_start.date()} to {prev_q_end.date()})"
    )
    ws_summary["A28"].font = Font(bold=True)
    write_dataframe(
        ws_summary,
        qcmp_df,
        start_row=29,
        start_col=1,
        currency_cols=["Current Quarter", "Previous Quarter"],
        pct_cols=["Pct Change"],
    )

    qcmp_start = 30
    qcmp_end = 29 + len(qcmp_df)
    for row in range(qcmp_start, qcmp_end + 1):
        section = ws_summary.cell(row=row, column=1).value
        pct_val = ws_summary.cell(row=row, column=5).value
        if pct_val is None:
            continue
        if section == "Income" and pct_val <= -0.10:
            for col in range(1, 6):
                ws_summary.cell(row=row, column=col).fill = RED_FILL
        if section == "Expense" and pct_val >= 0.10:
            for col in range(1, 6):
                ws_summary.cell(row=row, column=col).fill = RED_FILL

    write_dataframe(ws_pnl, pnl_df, currency_cols=[pnl_df.columns[-1]])
    write_dataframe(ws_bs, bs_df, currency_cols=[bs_df.columns[-1]])
    write_dataframe(ws_cf, cf_df, currency_cols=[cf_df.columns[-1]])

    for ws in [wb["Raw_GL"], ws_summary, ws_pnl, ws_bs, ws_cf]:
        auto_width(ws)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-ledger", required=True, help="Workbook containing Raw_GL sheet")
    parser.add_argument("--map", required=True, help="account_map.json")
    parser.add_argument("--stmt", required=True, help="Original bank statement CSV")
    parser.add_argument("--asof", required=False, help="Optional cutoff date YYYY-MM-DD")
    parser.add_argument("--xlsx", required=True, help="Output XLSX file path")
    args = parser.parse_args()

    ledger_path = Path(args.xlsx_ledger)
    output_path = Path(args.xlsx)

    df = load_transactions(ledger_path)
    mapping = load_account_map(args.map)
    stmt_header = parse_statement_header(args.stmt)
    df = apply_types(df, mapping)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(df, stmt_header, ledger_path, output_path, args.asof)
    print(f"Exported XLSX workbook: {output_path}")


if __name__ == "__main__":
    main()