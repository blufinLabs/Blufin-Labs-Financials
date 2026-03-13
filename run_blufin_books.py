#!/usr/bin/env python3

import csv
import hashlib
import json
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(".")

STATEMENT_DIR = ROOT / "Finance" / "statements"
REPORT_DIR = ROOT / "Finance" / "reports"
WORKBOOK = REPORT_DIR / "blufin_financial_statements.xlsx"

NEW_TX = ROOT / "new_transactions.csv"
RULES = ROOT / "categorization_rules.json"
ACCOUNT_MAP = ROOT / "account_map.json"
PROCESSED_LOG = ROOT / "processed_statements.json"


def file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while chunk := f.read(8192):
            h.update(chunk)
    return h.hexdigest()


def load_processed():
    if not PROCESSED_LOG.exists():
        return []
    with open(PROCESSED_LOG, "r", encoding="utf-8") as f:
        return json.load(f)


def save_processed(processed):
    with open(PROCESSED_LOG, "w", encoding="utf-8") as f:
        json.dump(processed, f, indent=2)


def is_statement_csv(path: Path) -> bool:
    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            lines = f.read().splitlines()
        return any(line.startswith("Date,Description,Amount,Running Bal.") for line in lines[:50])
    except Exception:
        return False


def discover_statements():
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    STATEMENT_DIR.mkdir(parents=True, exist_ok=True)
    candidates = sorted(STATEMENT_DIR.glob("*.csv"))
    return [p for p in candidates if is_statement_csv(p)]


def tx_key(row):
    s = f"{row['date']}|{row['description']}|{row['amount']}"
    return hashlib.md5(s.encode()).hexdigest()


def ensure_workbook_with_raw_gl():
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    if WORKBOOK.exists():
        wb = openpyxl.load_workbook(WORKBOOK)
        if "Raw_GL" not in wb.sheetnames:
            ws = wb.create_sheet("Raw_GL")
            ws.append(["date", "description", "amount", "category", "counterparty", "account", "memo"])
            wb.save(WORKBOOK)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw_GL"
    ws.append(["date", "description", "amount", "category", "counterparty", "account", "memo"])
    wb.save(WORKBOOK)


def read_raw_gl():
    ensure_workbook_with_raw_gl()
    df = pd.read_excel(WORKBOOK, sheet_name="Raw_GL")

    expected = ["date", "description", "amount", "category", "counterparty", "account", "memo"]
    for col in expected:
        if col not in df.columns:
            df[col] = ""

    df = df[expected].copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df["description"] = df["description"].fillna("").astype(str)
    df["category"] = df["category"].fillna("").astype(str)
    df["counterparty"] = df["counterparty"].fillna("").astype(str)
    df["account"] = df["account"].fillna("").astype(str)
    df["memo"] = df["memo"].fillna("").astype(str)

    df = df[df["date"].notna() & df["amount"].notna()].copy()
    df["date"] = df["date"].dt.strftime("%Y-%m-%d")

    return df


def write_raw_gl(df: pd.DataFrame):
    ensure_workbook_with_raw_gl()

    wb = openpyxl.load_workbook(WORKBOOK)

    if "Raw_GL" in wb.sheetnames:
        ws = wb["Raw_GL"]
        wb.remove(ws)

    ws = wb.create_sheet("Raw_GL", 0)
    ws.append(["date", "description", "amount", "category", "counterparty", "account", "memo"])

    for _, row in df.iterrows():
        ws.append([
            row.get("date", ""),
            row.get("description", ""),
            row.get("amount", 0.0),
            row.get("category", ""),
            row.get("counterparty", ""),
            row.get("account", ""),
            row.get("memo", ""),
        ])

    wb.save(WORKBOOK)


def merge_into_raw_gl(new_path: Path):
    new_df = pd.read_csv(new_path)

    expected = ["date", "description", "amount", "category", "counterparty", "account", "memo"]
    for col in expected:
        if col not in new_df.columns:
            new_df[col] = ""

    new_df = new_df[expected].copy()
    new_df["date"] = pd.to_datetime(new_df["date"], errors="coerce")
    new_df["amount"] = pd.to_numeric(new_df["amount"], errors="coerce")
    new_df["description"] = new_df["description"].fillna("").astype(str)
    new_df["category"] = new_df["category"].fillna("").astype(str)
    new_df["counterparty"] = new_df["counterparty"].fillna("").astype(str)
    new_df["account"] = new_df["account"].fillna("").astype(str)
    new_df["memo"] = new_df["memo"].fillna("").astype(str)
    new_df = new_df[new_df["date"].notna() & new_df["amount"].notna()].copy()
    new_df["date"] = new_df["date"].dt.strftime("%Y-%m-%d")

    master_df = read_raw_gl()

    if master_df.empty:
        merged = new_df.copy()
    else:
        master_df["tx_key"] = master_df.apply(tx_key, axis=1)
        new_df["tx_key"] = new_df.apply(tx_key, axis=1)

        merged = pd.concat(
            [master_df, new_df[~new_df["tx_key"].isin(master_df["tx_key"])]],
            ignore_index=True
        )
        merged = merged.drop(columns=["tx_key"], errors="ignore")

    merged = merged.sort_values(["date", "description", "amount"]).reset_index(drop=True)
    write_raw_gl(merged)

    print(f"Raw_GL now contains {len(merged)} transactions.")


def run_import(stmt: Path):
    subprocess.run(
        [
            "python",
            "import_stmt.py",
            "--input", str(stmt),
            "--output", str(NEW_TX),
            "--rules", str(RULES),
        ],
        check=True,
    )


def run_review(stmt: Path):
    subprocess.run(
        [
            "python",
            "review_transactions.py",
            "--input", str(NEW_TX),
            "--output", str(NEW_TX),
            "--rules", str(RULES),
            "--map", str(ACCOUNT_MAP),
            "--stmt", str(stmt),
        ],
        check=True,
    )


def run_reports(stmt: Path):
    subprocess.run(
        [
            "python",
            "blufin_accounting_engine.py",
            "--xlsx-ledger", str(WORKBOOK),
            "--map", str(ACCOUNT_MAP),
            "--stmt", str(stmt),
            "--xlsx", str(WORKBOOK),
        ],
        check=True,
    )


def auto_year_close():
    return


def main():
    ensure_workbook_with_raw_gl()

    processed = load_processed()
    statements = discover_statements()

    if not statements:
        print("No statements found.")
        return

    new_statements = []
    for stmt in statements:
        h = file_hash(stmt)
        if not any(p["hash"] == h for p in processed):
            new_statements.append((stmt, h))

    if not new_statements:
        print("No new statements detected.")
        return

    for stmt, h in new_statements:
        print("\n=====================================")
        print(f"Processing statement: {stmt.name}")
        print("=====================================")

        print("\nImporting statement")
        run_import(stmt)

        print("\nReviewing transactions")
        run_review(stmt)

        print("\nMerging transactions into Raw_GL")
        merge_into_raw_gl(NEW_TX)

        processed.append({
            "file": stmt.name,
            "hash": h
        })
        save_processed(processed)

    auto_year_close()

    print("\nGenerating financial reports")
    run_reports(new_statements[-1][0])

    print("\nAll statements processed.")
    print(f"Workbook updated: {WORKBOOK}")


if __name__ == "__main__":
    main()